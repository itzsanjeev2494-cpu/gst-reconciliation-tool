import pandas as pd
import os
import shutil
import sys
from datetime import date
from openpyxl import load_workbook


def find_column(df, possible_names):
    for col in df.columns:
        for name in possible_names:
            if str(col).strip().lower() == name.lower():
                return col
    return None


def clear_sheet(ws, start_row):
    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(
        min_row=start_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col
    ):
        for cell in row:
            cell.value = None


def reconcile_data(purchase_file, gstr2b_file):
    purchase_df = pd.read_excel(purchase_file)
    gstr2b_df = pd.read_excel(gstr2b_file)

    # Detect columns dynamically
    purchase_tax_col = find_column(
        purchase_df,
        ["Taxable Amt", "Taxable Amount", "Taxable Value"]
    )

    purchase_gst_col = find_column(
        purchase_df,
        ["GST", "GST Amount", "GST Amt", "GSTR Amount"]
    )

    gstr2b_tax_col = find_column(
        gstr2b_df,
        ["Taxable Amt", "Taxable Amount", "Taxable Value"]
    )

    gstr2b_gst_col = find_column(
        gstr2b_df,
        ["GST", "GST Amount", "GST Amt", "GSTR Amount"]
    )

    # Rename columns
    purchase_df.rename(columns={
        purchase_tax_col: "Taxable Amt (Books)",
        purchase_gst_col: "GST (Books)"
    }, inplace=True)

    gstr2b_df.rename(columns={
        gstr2b_tax_col: "Taxable Amt (2A)",
        gstr2b_gst_col: "GST (2A)"
    }, inplace=True)

    # Full merge for all scenarios
    merged = pd.merge(
        purchase_df,
        gstr2b_df[
            ["Invoice No", "Taxable Amt (2A)", "GST (2A)"]
        ],
        on="Invoice No",
        how="outer"
    )

    # Fill blanks
    merged["Taxable Amt (Books)"] = merged["Taxable Amt (Books)"].fillna(0)
    merged["GST (Books)"] = merged["GST (Books)"].fillna(0)
    merged["Taxable Amt (2A)"] = merged["Taxable Amt (2A)"].fillna(0)
    merged["GST (2A)"] = merged["GST (2A)"].fillna(0)

    # Calculate differences
    merged["Tax Diff (₹)"] = (
        merged["Taxable Amt (Books)"] - merged["Taxable Amt (2A)"]
    )

    merged["GST Diff (₹)"] = (
        merged["GST (Books)"] - merged["GST (2A)"]
    )

    # Status logic
    def get_status(row):
        if row["Taxable Amt (Books)"] > 0 and row["Taxable Amt (2A)"] > 0:
            if abs(row["Tax Diff (₹)"]) <= 1 and abs(row["GST Diff (₹)"]) <= 1:
                return "Matched"
            return "Mismatch"

        elif row["Taxable Amt (Books)"] > 0 and row["Taxable Amt (2A)"] == 0:
            return "Missing in 2A"

        elif row["Taxable Amt (Books)"] == 0 and row["Taxable Amt (2A)"] > 0:
            return "Extra in 2A"

    merged["Status"] = merged.apply(get_status, axis=1)

    mismatches = merged[merged["Status"] == "Mismatch"].copy()

    itc_risk = merged[merged["Status"] == "Missing in 2A"].copy()
    itc_risk["Taxable Amt (₹)"] = itc_risk["Taxable Amt (Books)"]
    itc_risk["ITC at Risk (₹)"] = itc_risk["GST (Books)"]
    itc_risk["Recommended Action"] = "Only in Books"

    # Summary
    summary = {
        "Total invoices (Books)": (
            len(purchase_df),
            purchase_df["Taxable Amt (Books)"].sum()
        ),
        "Matched": (
            len(merged[merged["Status"] == "Matched"]),
            merged[merged["Status"] == "Matched"]["Taxable Amt (Books)"].sum()
        ),
        "Mismatches": (
            len(mismatches),
            mismatches["Taxable Amt (Books)"].sum()
        ),
        "Missing in 2A": (
            len(itc_risk),
            itc_risk["Taxable Amt (Books)"].sum()
        ),
        "Extra in 2A": (
            len(merged[merged["Status"] == "Extra in 2A"]),
            merged[merged["Status"] == "Extra in 2A"]["Taxable Amt (2A)"].sum()
        ),
        "ITC safe to claim": (
            "-",
            merged[merged["Status"] == "Matched"]["GST (Books)"].sum()
        ),
        "ITC at risk": (
            "-",
            itc_risk["GST (Books)"].sum()
        )
    }

    return {
        "summary": summary,
        "full_reconciliation": merged,
        "mismatches": mismatches,
        "itc_risk": itc_risk
    }


def save_report(result, tax_month):
    template_file = "GST_Reconciliation_Report_April (3).xlsx"

    os.makedirs("reports", exist_ok=True)

    output_file = f"reports/GST_Reconciliation_Report_{tax_month}.xlsx"

    shutil.copy(template_file, output_file)

    wb = load_workbook(output_file)

    # Summary Sheet
    ws = wb["Summary"]
    clear_sheet(ws, 5)

    ws["A2"] = f"Generated on {date.today().strftime('%d %b %Y')} | Tolerance: ₹1.0"

    row_num = 5
    for metric, values in result["summary"].items():
        ws.cell(row=row_num, column=1, value=metric)
        ws.cell(row=row_num, column=2, value=values[0])
        ws.cell(row=row_num, column=3, value=values[1])
        row_num += 1

    # Full Reconciliation
    ws = wb["Full Reconciliation"]
    clear_sheet(ws, 3)

    start_row = 3
    for _, row in result["full_reconciliation"].iterrows():
        for col_idx, value in enumerate(row.tolist(), start=1):
            ws.cell(row=start_row, column=col_idx, value=value)
        start_row += 1

    # Mismatches
    ws = wb["Mismatches"]
    clear_sheet(ws, 3)

    start_row = 3
    for _, row in result["mismatches"].iterrows():
        for col_idx, value in enumerate(row.tolist(), start=1):
            ws.cell(row=start_row, column=col_idx, value=value)
        start_row += 1

    # ITC at Risk
    ws = wb["ITC at Risk"]
    clear_sheet(ws, 4)

    start_row = 4
    for _, row in result["itc_risk"].iterrows():
        ws.cell(start_row, 1, row["Invoice No"])
        ws.cell(start_row, 2, row["Vendor Name"])
        ws.cell(start_row, 3, row["GSTIN"])
        ws.cell(start_row, 4, row["Invoice Date"])
        ws.cell(start_row, 5, row["Taxable Amt (₹)"])
        ws.cell(start_row, 6, row["ITC at Risk (₹)"])
        ws.cell(start_row, 7, row["Recommended Action"])
        start_row += 1

    wb.save(output_file)

    return output_file


def main():
    if len(sys.argv) != 4:
        print("Usage: python gst_reconciliation.py purchase.xlsx gstr2b.xlsx month")
        return

    purchase_file = sys.argv[1]
    gstr2b_file = sys.argv[2]
    tax_month = sys.argv[3]

    print("Loading data...")
    print("Reconciling...")

    result = reconcile_data(
        purchase_file,
        gstr2b_file
    )

    print("Saving report...")

    output_path = save_report(
        result,
        tax_month
    )

    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()